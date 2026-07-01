import hashlib
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = Path(r"c:\Users\miraveta\Downloads\AyuGram Desktop\методичка по меню.xlsx")
MD = Path(
    r"C:\Users\miraveta\.cursor\projects\c-Users-miraveta-Pictures-chexov\uploads\restaurant-0.md"
)
OUT = ROOT / "src" / "data" / "menu.json"


def normalize(name: str) -> str:
    name = name.lower().strip()
    name = name.replace("ё", "е")
    name = re.sub(r"\s+", " ", name)
    name = re.sub(r'[«»"()]', "", name)
    return name


def make_id(name: str, category: str) -> str:
    raw = f"{category}:{name}".encode("utf-8")
    return hashlib.md5(raw).hexdigest()[:12]


def parse_xlsx() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    items: list[dict] = []
    category = ""

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = [str(c).strip() if c is not None else "" for c in row]
        name, desc, comp, allerg = vals[1], vals[2], vals[3], vals[4]

        if name in ("Название", ""):
            continue

        if desc == "" and comp == "" and allerg == "" and name:
            category = name
            continue

        if name and category:
            items.append(
                {
                    "name": name,
                    "category": category,
                    "description": desc,
                    "composition": comp,
                    "allergens": allerg,
                }
            )

    return items


def parse_prices() -> dict[str, int]:
    md = MD.read_text(encoding="utf-8")
    price_map: dict[str, int] = {}

    for block in md.split(" подробнее "):
        match = re.search(r"(\d[\d\s]*)\s*₽\s*$", block.strip())
        if not match:
            continue

        price = int(match.group(1).replace(" ", ""))
        before = block[: match.start()].strip()
        weight_match = re.search(r"\s+(\d+\s*(?:г|мл))\s+", before)

        if weight_match:
            name = before[: weight_match.start()].strip()
        else:
            name = before.split("  ")[0].strip()

        name = re.sub(r"\s+", " ", name)
        if len(name) > 2 and "18+" not in name and "авторизуйтесь" not in name.lower():
            price_map[normalize(name)] = price

    return price_map


def find_price(name: str, price_map: dict[str, int]) -> int:
    key = normalize(name)
    if key in price_map:
        return price_map[key]

    for price_key, price in price_map.items():
        if price_key.startswith(key[:18]) or key.startswith(price_key[:18]):
            return price

    return 0


def main() -> None:
    items = parse_xlsx()
    price_map = parse_prices()
    merged = []

    for item in items:
        price = find_price(item["name"], price_map)
        merged.append(
            {
                **item,
                "price": price,
                "id": make_id(item["name"], item["category"]),
            }
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(merged, ensure_ascii=False, indent=2), encoding="utf-8")

    with_price = sum(1 for x in merged if x["price"] > 0)
    print(f"Saved {len(merged)} items ({with_price} with prices) -> {OUT}")


if __name__ == "__main__":
    main()
